import streamlit as st
import pandas as pd
import pandasql as ps
import json
import os
import requests
import threading
import time
import base64
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from dotenv import load_dotenv
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global variables for data management
DATA_FILE = "oih_database.json"
API_URL = os.getenv("OIH_API_KEY")
UPDATE_INTERVAL = 20 * 60  # 20 minutes in seconds

# Configuration for large dataset handling
CONFIG = {
    'MAX_ROWS_FOR_LLM': 50,      # Maximum rows to send to LLM for response generation
    'MAX_CHARS_FOR_LLM': 2000,   # Maximum characters to send to LLM
    'LARGE_RESULT_THRESHOLD': 100, # Threshold for "large" results in UI
    'PREVIEW_ROWS': 10,          # Number of rows to show in preview
    'MAX_SAMPLE_ROWS': 10        # Maximum rows for LLM analysis
}

# Load environment variables
load_dotenv()

def load_data_from_file():
    """Load data from JSON file"""
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r") as f:
                json_data = json.load(f)
            df = pd.DataFrame(json_data)
            logger.info(f"Data loaded successfully. Shape: {df.shape}")
            return df, datetime.now()
        else:
            logger.warning(f"File {DATA_FILE} not found")
            return pd.DataFrame(), None
    except Exception as e:
        logger.error(f"Error loading data: {str(e)}")
        return pd.DataFrame(), None

def fetch_data_from_api():
    """Fetch data from API and save to JSON file"""
    if not API_URL:
        logger.warning("No API URL configured")
        return False
        
    try:
        response = requests.get(API_URL, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        data = data["ServiceRes"]
        
        with open(DATA_FILE, "w") as f:
            json.dump(data, f, indent=2)
        
        logger.info(f"Data updated from API. Records: {len(data) if isinstance(data, list) else 'N/A'}")
        return True
        
    except Exception as e:
        logger.error(f"Error fetching from API: {str(e)}")
        return False

def should_update_data(last_update):
    """Check if data should be updated"""
    if not last_update:
        return True
    return datetime.now() - last_update > timedelta(seconds=UPDATE_INTERVAL)

def get_current_data():
    """Get current data, update if necessary"""
    # Check session state for cached data
    if 'df' not in st.session_state or 'last_update' not in st.session_state:
        df, last_update = load_data_from_file()
        st.session_state.df = df
        st.session_state.last_update = last_update
    
    # Check if we need to update
    if should_update_data(st.session_state.last_update):
        if API_URL and fetch_data_from_api():
            df, last_update = load_data_from_file()
            st.session_state.df = df
            st.session_state.last_update = last_update
    
    return st.session_state.df

def initialize_llm():
    """Initialize the language model"""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        st.error("OpenAI API key not found. Please set OPENAI_API_KEY in your .env file.")
        st.stop()
    
    return ChatOpenAI(model_name="gpt-3.5-turbo", temperature=0, api_key=api_key)

def create_sql_prompt():
    """Create the SQL generation prompt template"""
    return PromptTemplate(
        input_variables=["user_input", "columns", "sample_data"],
        template="""
You are a pandasql expert working with a DataFrame called 'df'.

Table Information:
- Columns: {columns}
- Sample data (first 3 rows): {sample_data}

Generate a **SQLite-compatible SQL query** that answers the user's question.
- Use tolerance column for date it is our main date.
- Use proper SQL syntax for pandasql
- Handle case-insensitive searches when appropriate
- Use proper aggregations and grouping
- Return only the SQL query, no explanation

User question: {user_input}
SQL query:"""
    )

def create_response_prompt():
    """Create the response generation prompt template"""
    return PromptTemplate(
        input_variables=["user_input", "result"],
        template="""
Generate a user-friendly answer based on the user's question and query result.
Make it conversational and informative.

User question: {user_input}
Query result: {result}

Provide a clear, concise answer:"""
    )

def create_chart_analysis_prompt():
    """Create prompt for chart analysis and generation"""
    return PromptTemplate(
        input_variables=["user_input", "result", "columns"],
        template="""
Analyze the user's question and query result to determine if a chart should be generated.

User question: {user_input}
Query result columns: {columns}
Data sample: {result}

Based on the user's question, determine:
1. Should a chart be generated? (yes/no)
2. What type of chart? (bar, line, pie, scatter, area)
3. What should be the x-axis column?
4. What should be the y-axis column?
5. What should be the chart title?
6. Should comparison data be added? (yes/no)
7. If comparison needed, what period should be compared?

Respond in this exact format:
GENERATE_CHART: yes/no
CHART_TYPE: type
X_AXIS: column_name
Y_AXIS: column_name
TITLE: chart title
COMPARISON: yes/no
COMPARISON_PERIOD: period description"""
    )

def generate_sql_query(user_input, df, llm):
    """Generate SQL query from user input"""
    sql_prompt = create_sql_prompt()
    sql_chain = LLMChain(llm=llm, prompt=sql_prompt)
    
    columns = ", ".join(df.columns)
    sample_data = df.head(3).to_string() if not df.empty else "No data available"
    
    return sql_chain.run({
        "user_input": user_input,
        "columns": columns,
        "sample_data": sample_data
    })

def analyze_for_chart_generation(user_input, result, llm):
    """Analyze if chart should be generated and what type"""
    try:
        if result.empty or len(result) == 0:
            return None
            
        chart_prompt = create_chart_analysis_prompt()
        chart_chain = LLMChain(llm=llm, prompt=chart_prompt)
        
        columns = ", ".join(result.columns)
        sample_data = result.head(3).to_string() if len(result) > 0 else "No data"
        
        analysis = chart_chain.run({
            "user_input": user_input,
            "result": sample_data,
            "columns": columns
        })
        
        # Parse the analysis
        lines = analysis.strip().split('\n')
        chart_config = {}
        
        for line in lines:
            if ':' in line:
                key, value = line.split(':', 1)
                chart_config[key.strip()] = value.strip()
        
        if chart_config.get('GENERATE_CHART', '').lower() == 'yes':
            return chart_config
        else:
            return None
            
    except Exception as e:
        logger.error(f"Error in chart analysis: {str(e)}")
        return None

def create_comparison_data(df, original_result, user_input, chart_config):
    """Create comparison data based on the original query and chart configuration"""
    try:
        comparison_data = None
        
        # Extract year and month patterns from user input
        user_lower = user_input.lower()
        
        if 'year' in user_lower and '2025' in user_input:
            # Compare with 2024
            comparison_query = f"""
            SELECT 'Year 2024' as period, SUM(CAST(shipped_value_usd as FLOAT)) as total_value
            FROM df 
            WHERE strftime('%Y', tolerance) = '2024'
            UNION ALL
            SELECT 'Year 2025' as period, SUM(CAST(shipped_value_usd as FLOAT)) as total_value
            FROM df 
            WHERE strftime('%Y', tolerance) = '2025'
            """
            try:
                comparison_data = ps.sqldf(comparison_query, {"df": df})
            except:
                pass
                
        elif 'month' in user_lower and 'january' in user_lower:
            # Compare January with February
            comparison_query = f"""
            SELECT 'January 2025' as period, SUM(CAST(shipped_value_usd as FLOAT)) as total_value
            FROM df 
            WHERE strftime('%Y-%m', tolerance) = '2025-01'
            UNION ALL
            SELECT 'February 2025' as period, SUM(CAST(shipped_value_usd as FLOAT)) as total_value
            FROM df 
            WHERE strftime('%Y-%m', tolerance) = '2025-02'
            """
            try:
                comparison_data = ps.sqldf(comparison_query, {"df": df})
            except:
                pass
        
        return comparison_data
        
    except Exception as e:
        logger.error(f"Error creating comparison data: {str(e)}")
        return None

def generate_chart(result_df, chart_config, comparison_data=None, user_input=""):
    """Generate appropriate chart based on configuration"""
    try:
        chart_type = chart_config.get('CHART_TYPE', 'bar').lower()
        title = chart_config.get('TITLE', 'Data Visualization')
        x_axis = chart_config.get('X_AXIS', '')
        y_axis = chart_config.get('Y_AXIS', '')
        
        # Use comparison data if available, otherwise use original result
        plot_data = comparison_data if comparison_data is not None and not comparison_data.empty else result_df
        
        if plot_data.empty:
            return None
            
        # Auto-detect columns if not specified
        if not x_axis or x_axis not in plot_data.columns:
            x_axis = plot_data.columns[0]
        if not y_axis or y_axis not in plot_data.columns:
            numeric_cols = plot_data.select_dtypes(include=['number']).columns
            y_axis = numeric_cols[0] if len(numeric_cols) > 0 else plot_data.columns[-1]
        
        # Create the chart based on type
        if chart_type == 'bar':
            fig = px.bar(
                plot_data, 
                x=x_axis, 
                y=y_axis, 
                title=title,
                color_discrete_sequence=['#1f77b4']
            )
        elif chart_type == 'line':
            fig = px.line(
                plot_data, 
                x=x_axis, 
                y=y_axis, 
                title=title,
                markers=True
            )
        elif chart_type == 'pie':
            fig = px.pie(
                plot_data, 
                values=y_axis, 
                names=x_axis, 
                title=title
            )
        elif chart_type == 'area':
            fig = px.area(
                plot_data, 
                x=x_axis, 
                y=y_axis, 
                title=title
            )
        else:  # default to bar
            fig = px.bar(
                plot_data, 
                x=x_axis, 
                y=y_axis, 
                title=title,
                color_discrete_sequence=['#1f77b4']
            )
        
        # Update layout
        fig.update_layout(
            height=400,
            showlegend=True,
            template="plotly_white",
            title_x=0.5,
            font=dict(size=12)
        )
        
        # Format y-axis for currency if it contains value/amount
        if 'value' in y_axis.lower() or 'amount' in y_axis.lower() or 'usd' in y_axis.lower():
            fig.update_yaxes(tickformat='$,.0f')
        
        return fig
        
    except Exception as e:
        logger.error(f"Error generating chart: {str(e)}")
        return None

def generate_user_response(user_input, result, llm):
    """Generate user-friendly response - optimized for large datasets"""
    
    if result.empty:
        return "No results found for your query."
    
    # Check if result is too large to send to LLM
    result_str = result.to_string()
    is_large_result = (len(result) > CONFIG['MAX_ROWS_FOR_LLM'] or 
                      len(result_str) > CONFIG['MAX_CHARS_FOR_LLM'])
    
    if is_large_result:
        # For large results, provide a summary without LLM
        total_rows = len(result)
        total_cols = len(result.columns)
        
        # Get basic statistics
        summary_parts = [
            f"Found {total_rows:,} records with {total_cols} columns."
        ]
        
        # Add column info
        if total_cols <= 10:  # Only show column names if not too many
            summary_parts.append(f"Columns: {', '.join(result.columns)}")
        
        # Add basic stats for numeric columns
        numeric_cols = result.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            summary_parts.append("The complete results are displayed in the table below.")
        
        return " ".join(summary_parts)
    
    else:
        # For smaller results, use LLM to generate friendly response
        response_prompt = create_response_prompt()
        response_chain = LLMChain(llm=llm, prompt=response_prompt)
        
        # Limit the data sent to LLM
        limited_result = result.head(CONFIG['MAX_SAMPLE_ROWS']) if len(result) > CONFIG['MAX_SAMPLE_ROWS'] else result
        limited_result_str = limited_result.to_string()
        
        try:
            response = response_chain.run({
                "user_input": user_input,
                "result": limited_result_str
            })
            
            # If original result was truncated, mention it
            if len(result) > CONFIG['MAX_SAMPLE_ROWS']:
                response += f"\n\n(Showing analysis of first {CONFIG['MAX_SAMPLE_ROWS']} rows. Total results: {len(result)} rows)"
            
            return response
            
        except Exception as e:
            logger.error(f"Error generating LLM response: {str(e)}")
            # Fallback to simple response
            return f"Query executed successfully. Found {len(result)} records. Results are displayed below."

def execute_query(sql_query, df):
    """Execute SQL query using pandasql"""
    try:
        result = ps.sqldf(sql_query, {"df": df})
        return result, None
    except Exception as e:
        return None, str(e)


#? __________________________________________ MAIN _________________________
# def generate_oih_report_data(df):
#     """Generate OIH report data using pandas DataFrame manipulation"""
#     try:
#         if df.empty:
#             return None, "No data available for OIH report"
        
#         # Ensure we have the required columns (adjust based on your actual column names)
#         required_cols = ['VenderName', 'customerName', 'POCreation', 'ReportData', 'ShippedValueInUSD', 'InHandqty']
        
#         # Check if required columns exist (adjust column names as needed)
#         available_cols = df.columns.tolist()
#         logger.info(f"Available columns: {available_cols}")
        
#         # Create a sample OIH report structure - adjust this based on your actual data structure
#         try:
#             # Group by vendor and customer
#             if 'VenderName' in df.columns and 'customerName' in df.columns:
#                 # Create monthly aggregations
#                 report_data = []
                
#                 # Get unique vendor-customer combinations
#                 vendor_customer_combinations = df[['VenderName', 'customerName']].drop_duplicates()
                
#                 for _, row in vendor_customer_combinations.iterrows():
#                     vendor = row['VenderName']
#                     customer = row['customerName']
                    
#                     # Filter data for this combination
#                     combo_data = df[(df['VenderName'] == vendor) & (df['customerName'] == customer)]
                    
#                     # Initialize row data
#                     row_data = {
#                         'VenderName': vendor,
#                         'customerName': customer
#                     }
                    
#                     # Calculate monthly values (adjust column names as needed)
#                     months = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
#                              'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
                    
#                     total_shipped = 0
#                     total_oih = 0
                    
#                     for month in months:
#                         # Shipped values
#                         shipped_col = f'ShippedValueInUSD_{month}'
#                         oih_col = f'OIH_{month}'
                        
#                         # Sample calculation - adjust based on your data structure
#                         shipped_value = combo_data.get('ShippedValueInUSD', pd.Series([0])).sum() / 12  # Sample division
#                         oih_value = combo_data.get('InHandqty', pd.Series([0])).sum() / 12  # Sample division
                        
#                         row_data[shipped_col] = shipped_value
#                         row_data[oih_col] = oih_value
                        
#                         total_shipped += shipped_value
#                         total_oih += oih_value
                    
#                     row_data['TotalShiped'] = total_shipped
#                     row_data['TotalOIH'] = total_oih
                    
#                     report_data.append(row_data)
                
#                 # Create DataFrame
#                 report_df = pd.DataFrame(report_data)
                
#                 # Convert numeric columns
#                 numeric_columns = [col for col in report_df.columns if col not in ['VenderName', 'customerName']]
#                 for col in numeric_columns:
#                     report_df[col] = pd.to_numeric(report_df[col], errors='coerce').fillna(0)
                
#                 return report_df, None
                
#             else:
#                 return  None
                
#         except Exception as e:
#             logger.error(f"Error in OIH report generation: {str(e)}")
#             return None, f"Error generating OIH report: {str(e)}"
            
#     except Exception as e:
#         logger.error(f"Error in generate_oih_report_data: {str(e)}")
#         return None, f"Error generating OIH report: {str(e)}"


# def generate_oih_report_data(df):
#     """Generate OIH report data using pandas DataFrame manipulation"""
#     try:
#         if df.empty:
#             return None, "No data available for OIH report"
        
#         # Ensure we have the required columns (adjust based on your actual column names)
#         required_cols = ['VenderName', 'CustomerName', 'ShippedDate', 'InHandqty', 'ShippedValueInUSD']
        
#         # Check if required columns exist (adjust column names as needed)
#         available_cols = df.columns.tolist()
#         logger.info(f"Available columns: {available_cols}")
        
#         # Convert 'ShippedDate' to datetime for accurate filtering
#         df['ShippedDate'] = pd.to_datetime(df['ShippedDate'], errors='coerce')

#         # Filter data for the year 2025 only
#         df_filtered = df[df['ShippedDate'].dt.year == 2025]
        
#         if df_filtered.empty:
#             return None, "No data available for the specified ShippedDate in 2025"

#         # Create a sample OIH report structure - adjust this based on your actual data structure
#         try:
#             # Group by vendor and customer
#             if 'VenderName' in df.columns and 'CustomerName' in df.columns:
#                 # Create monthly aggregations
#                 report_data = []
                
#                 # Get unique vendor-customer combinations
#                 vendor_customer_combinations = df_filtered[['VenderName', 'CustomerName']].drop_duplicates()
                
#                 for _, row in vendor_customer_combinations.iterrows():
#                     vendor = row['VenderName']
#                     customer = row['CustomerName']
                    
#                     # Filter data for this combination
#                     combo_data = df_filtered[(df_filtered['VenderName'] == vendor) & (df_filtered['CustomerName'] == customer)]
                    
#                     # Initialize row data
#                     row_data = {
#                         'VenderName': vendor,
#                         'CustomerName': customer
#                     }
                    
#                     # Calculate monthly values (adjust column names as needed)
#                     months = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
#                              'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
                    
#                     total_shipped = 0
#                     total_oih = 0
                    
#                     for month in months:
#                         # Shipped values
#                         shipped_col = f'ShippedValueInUSD_{month}'
#                         oih_col = f'OIH_{month}'
                        
#                         # Filter data for the specific month
#                         shipped_value = combo_data[combo_data['ShippedDate'].dt.month == months.index(month) + 1]
#                         oih_value = shipped_value['InHandqty'].sum()  # Sum of 'InHandqty' for this month
                        
#                         row_data[shipped_col] = shipped_value['ShippedValueInUSD'].sum()  # Total ShippedValue for the month
#                         row_data[oih_col] = oih_value  # Total OIH for the month
                        
#                         total_shipped += row_data[shipped_col]
#                         total_oih += row_data[oih_col]
                    
#                     row_data['TotalShiped'] = total_shipped
#                     row_data['TotalOIH'] = total_oih
                    
#                     report_data.append(row_data)
                
#                 # Create DataFrame
#                 report_df = pd.DataFrame(report_data)
                
#                 # Convert numeric columns
#                 numeric_columns = [col for col in report_df.columns if col not in ['VenderName', 'CustomerName']]
#                 for col in numeric_columns:
#                     report_df[col] = pd.to_numeric(report_df[col], errors='coerce').fillna(0)
                
#                 return report_df, None
#             else:
#                 return None, "Missing required columns in data"
                
#         except Exception as e:
#             logger.error(f"Error in OIH report generation: {str(e)}")
#             return None, f"Error generating OIH report: {str(e)}"
            
#     except Exception as e:
#         logger.error(f"Error in generate_oih_report_data: {str(e)}")
#         return None, f"Error generating OIH report: {str(e)}"

def generate_oih_report_data(df):
    """Generate OIH report data using pandas DataFrame manipulation"""
    try:
        if df.empty:
            return None, "No data available for OIH report"
        
        # Ensure we have the required columns (adjust based on your actual column names)
        required_cols = ['VenderName', 'CustomerName', 'ShippedDate', 'ReportData', 'ShippedValueInUSD',"InHandValue"]
        
        # Check if required columns exist (adjust column names as needed)
        available_cols = df.columns.tolist()
        logger.info(f"Available columns: {available_cols}")
        
        # Convert 'ShippedDate' to datetime for accurate filtering
        df['ShippedDate'] = pd.to_datetime(df['ShippedDate'], errors='coerce')

        # Filter data for the year 2025 only
        df_filtered = df[df['ShippedDate'].dt.year == 2025]
        
        if df_filtered.empty:
            return None, "No data available for the specified ShippedDate in 2025"

        # Create a sample OIH report structure - adjust this based on your actual data structure
        try:
            # Group by vendor and customer
            if 'VenderName' in df.columns and 'CustomerName' in df.columns:
                # Create monthly aggregations
                report_data = []
                
                # Get unique vendor-customer combinations
                vendor_customer_combinations = df_filtered[['VenderName', 'CustomerName']].drop_duplicates()
                
                for _, row in vendor_customer_combinations.iterrows():
                    vendor = row['VenderName']
                    customer = row['CustomerName']
                    
                    # Filter data for this combination
                    combo_data = df_filtered[(df_filtered['VenderName'] == vendor) & (df_filtered['CustomerName'] == customer)]
                    
                    # Initialize row data
                    row_data = {
                        'VenderName': vendor,
                        'CustomerName': customer
                    }
                    
                    # Determine the column to use based on 'ReportData'
                    if 'ReportData' in combo_data.columns:
                        report_type = combo_data['ReportData'].iloc[0]  # Assuming 'ReportData' has the same value for the group
                    else:
                        report_type = None
                    
                    # Calculate monthly values (adjust column names as needed)
                    months = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                             'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
                    
                    total_shipped = 0
                    total_oih = 0
                    
                    for month in months:
                        # Shipped values
                        shipped_col = f'ShippedValueInUSD_{month}'
                        oih_col = f'OIH_{month}'
                        
                        # Filter data for the specific month
                        month_data = combo_data[combo_data['ShippedDate'].dt.month == months.index(month) + 1]
                        
                        if report_type == 'Shipped':
                            # Sum of 'ShippedValueInUSD' for this month
                            shipped_value = month_data['ShippedValueInUSD'].sum()
                            row_data[shipped_col] = shipped_value
                            total_shipped += shipped_value
                            row_data[oih_col] = 0  # No OIH value for "Shipped"
                        elif report_type == 'OIH':
                            # Sum of 'InHandValue' for this month
                            oih_value = month_data['InHandValue'].sum()
                            row_data[oih_col] = oih_value
                            total_oih += oih_value
                            row_data[shipped_col] = 0  # No shipped value for "OIH"
                        else:
                            # Handle other cases if needed
                            row_data[shipped_col] = 0
                            row_data[oih_col] = 0
                    
                    # Add totals
                    row_data['TotalShiped'] = total_shipped
                    row_data['TotalOIH'] = total_oih
                    
                    report_data.append(row_data)
                
                # Create DataFrame for the report
                report_df = pd.DataFrame(report_data)
                
                # Convert numeric columns
                numeric_columns = [col for col in report_df.columns if col not in ['VenderName', 'CustomerName']]
                for col in numeric_columns:
                    report_df[col] = pd.to_numeric(report_df[col], errors='coerce').fillna(0)
                
                return report_df, None
            else:
                return None, "Missing required columns in data"
                
        except Exception as e:
            logger.error(f"Error in OIH report generation: {str(e)}")
            return None, f"Error generating OIH report: {str(e)}"
            
    except Exception as e:
        logger.error(f"Error in generate_oih_report_data: {str(e)}")
        return None, f"Error generating OIH report: {str(e)}"


#? __________________________________________ MAIN____________________________________________-
# def create_oih_excel_report(report_df):
#     """Create Excel report from OIH data"""
#     try:
#         if report_df is None or report_df.empty:
#             return None
            
#         # Column mapping
#         column_mapping = {
#             "VenderName": "Merchandiser",
#             "customerName": "Customer Name",
#             "ShippedValueInUSD_JANUARY": "JAN - Shipped USD",
#             "ShippedValueInUSD_FEBRUARY": "FEB - Shipped USD",
#             "ShippedValueInUSD_MARCH": "MAR - Shipped USD",
#             "ShippedValueInUSD_APRIL": "APR - Shipped USD",
#             "ShippedValueInUSD_MAY": "MAY - Shipped USD",
#             "ShippedValueInUSD_JUNE": "JUN - Shipped USD",
#             "ShippedValueInUSD_JULY": "JUL - Shipped USD",
#             "ShippedValueInUSD_AUGUST": "AUG - Shipped USD",
#             "ShippedValueInUSD_SEPTEMBER": "SEP - Shipped USD",
#             "ShippedValueInUSD_OCTOBER": "OCT - Shipped USD",
#             "ShippedValueInUSD_NOVEMBER": "NOV - Shipped USD",
#             "ShippedValueInUSD_DECEMBER": "DEC - Shipped USD",
#             "OIH_JANUARY": "JAN - OIH",
#             "OIH_FEBRUARY": "FEB - OIH",
#             "OIH_MARCH": "MAR - OIH",
#             "OIH_APRIL": "APR - OIH",
#             "OIH_MAY": "MAY - OIH",
#             "OIH_JUNE": "JUN - OIH",
#             "OIH_JULY": "JUL - OIH",
#             "OIH_AUGUST": "AUG - OIH",
#             "OIH_SEPTEMBER": "SEP - OIH",
#             "OIH_OCTOBER": "OCT - OIH",
#             "OIH_NOVEMBER": "NOV - OIH",
#             "OIH_DECEMBER": "DEC - OIH",
#             "TotalShiped": "Total Shipped (USD)",
#             "TotalOIH": "Total OIH"
#         }
        
#         # Apply column mapping only for columns that exist
#         existing_mapping = {k: v for k, v in column_mapping.items() if k in report_df.columns}
#         report_df_renamed = report_df.rename(columns=existing_mapping)
        

        
#         # Create Excel workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "OIH Report"
        
#         # Define styles
#         banner_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
#         banner_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
#         banner_alignment = Alignment(horizontal="center", vertical="center")
        
#         header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
#         header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
#         header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
#         left_alignment = Alignment(horizontal="left", vertical="center", indent=1)
#         center_alignment = Alignment(horizontal="center", vertical="center")
#         right_alignment = Alignment(horizontal="right", vertical="center", indent=1)
        
#         cell_font = Font(name="Calibri", size=10)
#         border_style = Border(
#             left=Side(style="thin"),
#             right=Side(style="thin"),
#             top=Side(style="thin"),
#             bottom=Side(style="thin"),
#         )
        
#         # Create banner header
#         last_col = get_column_letter(len(report_df_renamed.columns))
        
#         # Row 1: Company name
#         ws.merge_cells('A1:B1')
#         ws['A1'] = "Synergies Bangladesh"
#         ws['A1'].font = banner_font
#         ws['A1'].fill = banner_fill
#         ws['A1'].alignment = banner_alignment
#         ws['A1'].border = border_style
        
#         ws.merge_cells(f'C1:{last_col}1')
#         ws['C1'] = "Synergies Bangladesh"
#         ws['C1'].font = banner_font
#         ws['C1'].fill = banner_fill
#         ws['C1'].alignment = banner_alignment
        
#         # Row 2: SOIH
#         ws.merge_cells(f'A2:{last_col}2')
#         ws['A2'] = "SOIH"
#         ws['A2'].font = banner_font
#         ws['A2'].fill = banner_fill
#         ws['A2'].alignment = banner_alignment
        
#         # Row 3: Year
#         ws.merge_cells(f'A3:{last_col}3')
#         ws['A3'] = "2025"
#         ws['A3'].font = banner_font
#         ws['A3'].fill = banner_fill
#         ws['A3'].alignment = banner_alignment
        
#         # Headers in row 4
#         for c_idx, column_name in enumerate(report_df_renamed.columns, start=1):
#             cell = ws.cell(row=4, column=c_idx, value=column_name)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = header_alignment
#             cell.border = border_style
        
#         # Data rows starting from row 5
#         for r_idx, (index, row_data) in enumerate(report_df_renamed.iterrows(), start=5):
#             for c_idx, (column_name, value) in enumerate(zip(report_df_renamed.columns, row_data), start=1):
#                 cell = ws.cell(row=r_idx, column=c_idx)
                
#                 if isinstance(value, (int, float)) and any(keyword in column_name.upper() for keyword in ['USD', 'SHIPPED', 'OIH']):
#                     cell.value = float(value) if value != 0 else "-"
#                     if 'USD' in column_name.upper() or 'SHIPPED' in column_name.upper():
#                         cell.number_format = '"$"#,##0'
#                     elif 'OIH' in column_name.upper():
#                         cell.number_format = '"$"#,##0'
#                     else:
#                         cell.number_format = '#,##0'
#                     cell.alignment = right_alignment
#                 else:
#                     cell.value = str(value) if value is not None else ""
#                     if column_name in ['Merchandiser', 'Customer Name']:
#                         cell.alignment = left_alignment
#                     else:
#                         cell.alignment = center_alignment
                
#                 cell.font = cell_font
#                 cell.border = border_style
        
#         # Set column widths
#         column_widths = {
#             'A': 25,   # Merchandiser
#             'B': 35,   # Customer Name
#         }
        
#         for col_idx in range(3, len(report_df_renamed.columns) + 1):
#             col_letter = get_column_letter(col_idx)
#             column_widths[col_letter] = 14
        
#         for col_letter, width in column_widths.items():
#             ws.column_dimensions[col_letter].width = width
        
#         # Set row heights
#         ws.row_dimensions[1].height = 28
#         ws.row_dimensions[2].height = 25
#         ws.row_dimensions[3].height = 25
#         ws.row_dimensions[4].height = 35
        
#         for row in range(5, len(report_df_renamed) + 5):
#             ws.row_dimensions[row].height = 20
        
#         # Apply border to banner rows
#         for row in range(1, 4):
#             for col in range(1, len(report_df_renamed.columns) + 1):
#                 ws.cell(row=row, column=col).border = border_style
        
#         # Save to BytesIO and convert to base64
#         excel_buffer = BytesIO()
#         wb.save(excel_buffer)
#         excel_buffer.seek(0)
        
#         # Convert to base64
#         base64_data = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
#         excel_buffer.close()
        
#         return base64_data
        
#     except Exception as e:
#         logger.error(f"Error creating OIH Excel report: {str(e)}")
#         return None

def create_oih_excel_report(report_df):
    """Create Excel report from OIH data with proper column mapping and sequence"""
    try:
        if report_df is None or report_df.empty:
            return None
            
        # Column mapping - sequence will follow the order defined here
        column_mapping = {
            "VenderName": "Merchandiser",
            "CustomerName": "Customer Name",  # Fixed the key case
            "ShippedValueInUSD_JANUARY": "JAN - Shipped USD",
            "ShippedValueInUSD_FEBRUARY": "FEB - Shipped USD",
            "ShippedValueInUSD_MARCH": "MAR - Shipped USD",
            "ShippedValueInUSD_APRIL": "APR - Shipped USD",
            "ShippedValueInUSD_MAY": "MAY - Shipped USD",
            "ShippedValueInUSD_JUNE": "JUN - Shipped USD",
            "ShippedValueInUSD_JULY": "JUL - Shipped USD",
            "ShippedValueInUSD_AUGUST": "AUG - Shipped USD",
            "ShippedValueInUSD_SEPTEMBER": "SEP - Shipped USD",
            "ShippedValueInUSD_OCTOBER": "OCT - Shipped USD",
            "ShippedValueInUSD_NOVEMBER": "NOV - Shipped USD",
            "ShippedValueInUSD_DECEMBER": "DEC - Shipped USD",
            "OIH_JANUARY": "JAN - OIH",
            "OIH_FEBRUARY": "FEB - OIH",
            "OIH_MARCH": "MAR - OIH",
            "OIH_APRIL": "APR - OIH",
            "OIH_MAY": "MAY - OIH",
            "OIH_JUNE": "JUN - OIH",
            "OIH_JULY": "JUL - OIH",
            "OIH_AUGUST": "AUG - OIH",
            "OIH_SEPTEMBER": "SEP - OIH",
            "OIH_OCTOBER": "OCT - OIH",
            "OIH_NOVEMBER": "NOV - OIH",
            "OIH_DECEMBER": "DEC - OIH",
            "TotalShiped": "Total Shipped (USD)",
            "TotalOIH": "Total OIH"
        }
        
        # Use the order from column_mapping keys as the desired sequence
        existing_columns = [col for col in column_mapping.keys() if col in report_df.columns]
        
        # Reorder DataFrame columns according to desired sequence
        report_df_reordered = report_df[existing_columns].copy()
        
        # Apply column mapping
        existing_mapping = {k: v for k, v in column_mapping.items() if k in report_df_reordered.columns}
        report_df_renamed = report_df_reordered.rename(columns=existing_mapping)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "OIH Report"
        
        # Define styles
        banner_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        banner_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        banner_alignment = Alignment(horizontal="center", vertical="center")
        
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        left_alignment = Alignment(horizontal="left", vertical="center", indent=1)
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center", indent=1)
        
        cell_font = Font(name="Calibri", size=10)
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Create banner header
        last_col = get_column_letter(len(report_df_renamed.columns))
        
        # Row 1: Company name
        ws.merge_cells('A1:B1')
        ws['A1'] = "Synergies Bangladesh"
        ws['A1'].font = banner_font
        ws['A1'].fill = banner_fill
        ws['A1'].alignment = banner_alignment
        ws['A1'].border = border_style
        
        ws.merge_cells(f'C1:{last_col}1')
        ws['C1'] = "Synergies Bangladesh"
        ws['C1'].font = banner_font
        ws['C1'].fill = banner_fill
        ws['C1'].alignment = banner_alignment
        
        # Row 2: SOIH
        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = "SOIH"
        ws['A2'].font = banner_font
        ws['A2'].fill = banner_fill
        ws['A2'].alignment = banner_alignment
        
        # Row 3: Year
        ws.merge_cells(f'A3:{last_col}3')
        ws['A3'] = "2025"
        ws['A3'].font = banner_font
        ws['A3'].fill = banner_fill
        ws['A3'].alignment = banner_alignment
        
        # Headers in row 4
        for c_idx, column_name in enumerate(report_df_renamed.columns, start=1):
            cell = ws.cell(row=4, column=c_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Data rows starting from row 5
        for r_idx, (index, row_data) in enumerate(report_df_renamed.iterrows(), start=5):
            for c_idx, (column_name, value) in enumerate(zip(report_df_renamed.columns, row_data), start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                
                if isinstance(value, (int, float)) and any(keyword in column_name.upper() for keyword in ['USD', 'SHIPPED', 'OIH']):
                    cell.value = float(value) if value != 0 else "-"
                    if 'USD' in column_name.upper() or 'SHIPPED' in column_name.upper():
                        cell.number_format = '"$"#,##0'
                    elif 'OIH' in column_name.upper():
                        cell.number_format = '"$"#,##0'
                    else:
                        cell.number_format = '#,##0'
                    cell.alignment = right_alignment
                else:
                    cell.value = str(value) if value is not None else ""
                    if column_name in ['Merchandiser', 'Customer Name']:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment
                
                cell.font = cell_font
                cell.border = border_style
        
        # Set column widths
        column_widths = {
            'A': 25,   # Merchandiser
            'B': 35,   # Customer Name
        }
        
        for col_idx in range(3, len(report_df_renamed.columns) + 1):
            col_letter = get_column_letter(col_idx)
            column_widths[col_letter] = 14
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 35
        
        for row in range(5, len(report_df_renamed) + 5):
            ws.row_dimensions[row].height = 20
        
        # Apply border to banner rows
        for row in range(1, 4):
            for col in range(1, len(report_df_renamed.columns) + 1):
                ws.cell(row=row, column=col).border = border_style
        
        # Save to BytesIO and convert to base64
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Convert to base64
        base64_data = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
        excel_buffer.close()
        
        return base64_data
        
    except Exception as e:
        logger.error(f"Error creating OIH Excel report: {str(e)}")
        return None


def process_user_question(user_input, df, llm):
    """Process user question and return complete response with chart analysis"""
    try:
        if df.empty:
            return {
                "error": "No data available",
                "sql": None,
                "result": None,
                "response": "I don't have any data to query. Please check your database file.",
                "chart": None,
                "chart_config": None
            }
        
        # Check if this is an OIH report request
        if 'oih report' in user_input.lower():
            report_df, error = generate_oih_report_data(df)
            if error:
                return {
                    "error": error,
                    "sql": None,
                    "result": None,
                    "response": f"Error generating OIH report: {error}",
                    "chart": None,
                    "chart_config": None,
                    "oih_report": None
                }
            
            # Create Excel report
            excel_data = create_oih_excel_report(report_df)
            
            return {
                "error": None,
                "sql": "OIH Report Generation",
                "result": report_df,
                "response": f"OIH Report generated successfully with {len(report_df)} records. You can download the Excel report below.",
                "chart": None,
                "chart_config": None,
                "oih_report": excel_data
            }
        
        # Generate SQL
        sql_query = generate_sql_query(user_input, df, llm)
        
        # Execute SQL
        result, error = execute_query(sql_query, df)
        
        if error:
            return {
                "error": error,
                "sql": sql_query,
                "result": None,
                "response": f"I encountered an error executing the query: {error}",
                "chart": None,
                "chart_config": None
            }
        
        # Generate user-friendly response
        response = generate_user_response(user_input, result, llm)
        
        # Analyze for chart generation
        chart_config = analyze_for_chart_generation(user_input, result, llm)
        chart = None
        
        if chart_config:
            # Create comparison data if needed
            comparison_data = None
            if chart_config.get('COMPARISON', '').lower() == 'yes':
                comparison_data = create_comparison_data(df, result, user_input, chart_config)
            
            # Generate chart
            chart = generate_chart(result, chart_config, comparison_data, user_input)
        
        return {
            "error": None,
            "sql": sql_query,
            "result": result,
            "response": response,
            "chart": chart,
            "chart_config": chart_config
        }
        
    except Exception as e:
        logger.error(f"Error processing query: {str(e)}")
        return {
            "error": str(e),
            "sql": None,
            "result": None,
            "response": f"I encountered an error: {str(e)}",
            "chart": None,
            "chart_config": None
        }

def display_sidebar_info(df):
    """Display database information in sidebar"""
    with st.sidebar:
        st.header("Database Info")
        
        # Refresh button
        if st.button("🔄 Refresh Data"):
            if API_URL and fetch_data_from_api():
                df, last_update = load_data_from_file()
                st.session_state.df = df
                st.session_state.last_update = last_update
                st.success("Data refreshed!")
                st.rerun()
            else:
                df, last_update = load_data_from_file()
                st.session_state.df = df
                st.session_state.last_update = last_update
                st.info("Data reloaded from file")
                st.rerun()
        
        # Show data info
        if not df.empty:
            st.success(f"✅ {len(df)} records loaded")
            st.write(f"**Columns:** {', '.join(df.columns)}")
            if st.session_state.get('last_update'):
                st.write(f"**Last Update:** {st.session_state.last_update.strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Show sample data
            with st.expander("Sample Data"):
                st.dataframe(df.head())
        else:
            st.error("❌ No data loaded")
        
        # Clear chat button
        if st.button("🗑️ Clear Chat"):
            st.session_state.chat_history = []
            st.rerun()

def display_chat_history():
    """Display chat history with optimized display for large results and charts"""
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    
    for i, chat in enumerate(st.session_state.chat_history):
        with st.container():
            st.markdown(f"**You:** {chat['user']}")
            st.markdown(f"**Bot:** {chat['bot']}")
            
            # Display chart if available
            if chat.get('chart') is not None:
                st.plotly_chart(chat['chart'], use_container_width=True)
            
            # Display OIH report download if available
            if chat.get('oih_report'):
                st.download_button(
                    label="📥 Download OIH Report (Excel)",
                    data=base64.b64decode(chat['oih_report']),
                    file_name=f"OIH_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"oih_download_{i}"
                )
            
            if chat.get('sql') and chat.get('result') is not None:
                result_df = chat['result']
                
                # Determine how to display results based on size
                if not result_df.empty:
                    result_size = len(result_df)
                    
                    if result_size > CONFIG['LARGE_RESULT_THRESHOLD']:
                        # For large results, show summary first
                        with st.expander(f"📊 View Results & SQL #{i+1} ({result_size:,} rows)"):
                            st.code(chat['sql'], language='sql')
                            
                            # Show summary statistics
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Total Rows", f"{result_size:,}")
                            with col2:
                                st.metric("Columns", len(result_df.columns))
                            with col3:
                                if result_df.select_dtypes(include=['number']).columns.any():
                                    numeric_cols = len(result_df.select_dtypes(include=['number']).columns)
                                    st.metric("Numeric Columns", numeric_cols)
                            
                            # Show first and last few rows
                            st.subheader("📄 Data Preview")
                            
                            # First N rows
                            st.write(f"**First {CONFIG['PREVIEW_ROWS']} rows:**")
                            st.dataframe(result_df.head(CONFIG['PREVIEW_ROWS']), use_container_width=True)
                            
                            if result_size > (CONFIG['PREVIEW_ROWS'] * 2):
                                st.write(f"**Last {CONFIG['PREVIEW_ROWS']} rows:**")
                                st.dataframe(result_df.tail(CONFIG['PREVIEW_ROWS']), use_container_width=True)
                            
                            # Download option for large datasets
                            csv = result_df.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Full Results as CSV",
                                data=csv,
                                file_name=f"query_results_{i+1}.csv",
                                mime="text/csv",
                                key=f"csv_download_{i}"
                            )
                    else:
                        # For smaller results, show everything
                        with st.expander(f"View SQL & Results #{i+1} ({result_size} rows)"):
                            st.code(chat['sql'], language='sql')
                            st.dataframe(result_df, use_container_width=True)
                else:
                    with st.expander(f"View SQL & Results #{i+1}"):
                        st.code(chat['sql'], language='sql')
                        st.write("No results found")
            
            st.divider()

def display_example_queries():
    """Display example queries including OIH Report"""
    st.markdown("---")
    st.header("💡 Example Questions")
    
    examples = [
        "What is total shipped value of year 2024",
        "Show me total shipped quantity of month January",
        "Which buyer has the highest shipped quantity?",
        "What is total booked value in USD?",
        "OIH Report"  # Added OIH Report example
    ]
    
    cols = st.columns(3)  # Changed to 3 columns to accommodate 5 examples
    for i, example in enumerate(examples):
        with cols[i % 3]:
            if st.button(example, key=f"example_{i}", use_container_width=True):
                # Simulate user input
                st.session_state.example_clicked = example
                st.rerun()

def handle_example_query(llm, df):
    """Handle example query if clicked"""
    if 'example_clicked' in st.session_state:
        user_input = st.session_state.example_clicked
        del st.session_state.example_clicked
        
        with st.spinner("Processing example question..."):
            result = process_user_question(user_input, df, llm)
        
        # Add to chat history
        chat_entry = {
            "user": user_input,
            "bot": result["response"],
            "sql": result["sql"],
            "result": result["result"],
            "error": result["error"],
            "chart": result.get("chart"),
            "chart_config": result.get("chart_config"),
            "oih_report": result.get("oih_report")
        }
        st.session_state.chat_history.append(chat_entry)
        st.rerun()

def display_chart_info():
    """Display information about chart generation capabilities"""
    with st.expander("📈 Chart Generation Info"):
        st.markdown("""
        **Automatic Chart Generation:**
        - Charts are automatically generated for relevant queries
        - Supported chart types: Bar, Line, Pie, Area, Scatter
        - Comparison charts are created when appropriate
        
        **Examples that generate charts:**
        - "Total shipped value by year" → Generates bar chart with year comparison
        - "Monthly shipped quantity" → Generates line chart with month-to-month comparison
        - "Top customers by value" → Generates bar chart
        
        **Chart Features:**
        - Interactive Plotly charts
        - Automatic formatting for currency values
        - Responsive design
        - Export capabilities
        """)

def main():
    """Main application function"""
    # Page configuration
    st.set_page_config(
        page_title="Talk to Database",
        page_icon="🗣️",
        layout="wide"
    )
    
    st.title("🗣️ Talk to Database")
    st.subheader("OIH Database with Chart Generation & Reports")
    st.markdown("Ask questions about your data in natural language and get visual insights!")
    
    # Initialize components
    try:
        llm = initialize_llm()
        df = get_current_data()
        
        # Display sidebar
        display_sidebar_info(df)
        
        # Display chart info
        display_chart_info()
        
        # Handle example queries
        handle_example_query(llm, df)
        
        # Chat interface
        st.header("💬 Chat Interface")
        
        # Display chat history
        display_chat_history()
        
        # Chat input
        user_input = st.chat_input("Ask a question about your data...")
        
        if user_input:
            with st.spinner("Processing your question..."):
                result = process_user_question(user_input, df, llm)
            
            # Add to chat history
            chat_entry = {
                "user": user_input,
                "bot": result["response"],
                "sql": result["sql"],
                "result": result["result"],
                "error": result["error"],
                "chart": result.get("chart"),
                "chart_config": result.get("chart_config"),
                "oih_report": result.get("oih_report")
            }
            st.session_state.chat_history.append(chat_entry)
            st.rerun()
        
        # Display example queries
        display_example_queries()
        
    except Exception as e:
        st.error(f"Application error: {str(e)}")
        logger.error(f"Application error: {str(e)}")

# Background data updater function (optional)
def start_background_updater():
    """Start background data updater if API_URL is configured"""
    if not API_URL:
        return
    
    def update_loop():
        while True:
            time.sleep(UPDATE_INTERVAL)
            fetch_data_from_api()
    
    thread = threading.Thread(target=update_loop, daemon=True)
    thread.start()
    logger.info("Background data updater started")

if __name__ == "__main__":
    # Start background updater
    start_background_updater()
    
    # Run main app
    main()